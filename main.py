"""Модуль с основной логикой."""
import os
import sqlite3
from openpyxl import load_workbook, Workbook

from manage_db import (add_data_in_tables, annotate_data_in_tables,
                       annotate_data_previous_year, create_table)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

COMPRES_TUPLE = ('вв', 'акв', 'дэн')

DB_NAME = 'db.sqlite3'
DB_FILE = 'file:' + os.path.join(BASE_DIR, DB_NAME)

uktol = list()
vv = list()
akv = list()
den = list()
others = list()

kirillic_name_and_lists = {
    'уктол': uktol,
    'вв': vv,
    'акв': akv,
    'дэн': den,
    'прочее': others
}

table_name_and_data = {
    'uktol': uktol,
    'vv': vv,
    'akv': akv,
    'den': den,
    'others': others
}

percentage = {
    'вв': 0.41,
    'акв': 0.51,
    'дэн': 0.08,
}


def count_per_kilometers(args):
    """Получает на вход кортеж из условных пар значений количества
    неисправностей и пробега за расчетное время. Для каждой пары
    расчитывается соответствующее значение ед. на млн.км. пробега."""
    result = list()
    for index, value in enumerate(args):
        if index & 1:
            continue
        try:
            res = value / args[index + 1]
        except ZeroDivisionError:
            res = 0
        result.append(value)
        result.append(res)
    return tuple(result)


def add_data_in_db_func(filename):
    """Получает в качестве аргумента путь к файлу excel с данными.
    Производит считывание данных, создание талбиц в БД, если
    они еще не были созданы, и запись данных в таблицы БД."""
    wb = load_workbook(filename=filename, data_only=True)
    wb.active = 0
    sheet = wb.active

    for i in range(1, sheet.max_row):
        field = sheet['A' + str(i + 1)]
        type_of_machine = field.value.lower()
        if type_of_machine == 'уктол' or type_of_machine in COMPRES_TUPLE:
            target = kirillic_name_and_lists[type_of_machine]
        else:
            target = others

        result = [
            col[i].value for col in sheet.iter_cols(0, sheet.max_column)
        ]
        if type_of_machine in COMPRES_TUPLE:
            result[-1] = result[-1] * percentage[type_of_machine]

        target.append(tuple(result))

    con = sqlite3.connect(DB_FILE, uri=True)
    cur = con.cursor()

    create_table(*table_name_and_data.keys(), cur=cur)

    for table, data in table_name_and_data.items():
        add_data_in_tables(table, cur=cur, data=data)

    con.commit()
    con.close()


def analyse_func(current_month, current_year):
    """Получает на вход 'текущие' месяц и год - дату на которую необходимо
    подготовить отчетные таблицы в результирующем файле excel. Создает
    папку reports, если она еще не создана, аннотирует данные таблиц БД
    по неисправным деталям на 'текущую' дату. Производит дополнительные
    операции с БД для получения данных за аналогичные периоды прошлых
    лет. Полученный результат записывается в результирующий файл
    excel 'analysis' с разбивкой на отдельные книги по типу оборудования."""
    col_names = (
        'Оборудование',
        'Узел',
        f'Кол-во за {current_year}г. в ед.',
        f'Кол-во за {current_year}г. в ед./млн.км.',
        f'Кол-во за {current_year - 1}г. в ед.',
        f'Кол-во за {current_year - 1}г. в ед./млн.км.',
        f'Кол-во за {current_year - 2}г. в ед.',
        f'Кол-во за {current_year - 2}г. в ед./млн.км.',
        f'Кол-во за {current_year - 3}г. в ед.',
        f'Кол-во за {current_year - 3}г. в ед./млн.км.',
        f'Кол-во за {current_year - 4}г. в ед.',
        f'Кол-во за {current_year - 4}г. в ед./млн.км.'
    )

    con = sqlite3.connect(DB_FILE, uri=True)
    cur = con.cursor()

    ann_uktol, ann_vv, ann_akv, ann_den, ann_others = (
        annotate_data_in_tables(
            item, current_month, current_year, cur=cur
        )
        for item in table_name_and_data.keys()
    )

    result_tables = {
        'уктол': ann_uktol,
        'вв': ann_vv,
        'акв': ann_akv,
        'дэн': ann_den,
        'прочее': ann_others
    }

    result_wb = Workbook()

    os.makedirs('reports', exist_ok=True)
    dest_filename = 'reports/analysis.xlsx'

    del result_wb['Sheet']

    wb_sheet = result_wb.create_sheet(title='общая')
    wb_sheet.append(col_names)

    for name, value in result_tables.items():
        wb_sheet = result_wb.create_sheet(title=name)
        wb_sheet.append(col_names)

        for item in value:
            details = item[:2]
            if details[0].lower() == 'вв':
                details = ('vv', details[1])
            elif details[0].lower() == 'акв':
                details = ('akv', details[1])
            elif details[0].lower() == 'дэн':
                details = ('den', details[1])
            elif details[0].lower() == 'уктол':
                details = ('uktol', details[1])
            else:
                details = ('others', details[1])

            prev_data = annotate_data_previous_year(
                details, current_month, current_year, cur=cur
            )
            prev_year_data_with_kilo = count_per_kilometers(prev_data)
            item = item[:-1] + prev_year_data_with_kilo
            wb_sheet.append(item)
            result_wb['общая'].append(item)

    result_wb.save(filename=dest_filename)

    con.commit()
    con.close()
